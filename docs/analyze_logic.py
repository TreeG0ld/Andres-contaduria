import openpyxl

file_path = r"d:\andresContador\Andres-contaduria\archivos extraccion de  data\planoNominalogica.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

print("# Resumen de Lógica del Excel\n")

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"## Hoja: {sheet_name}\n")
    
    # Analyze header or assume structure
    # We noticed columns: F (Cuenta), V (Debito), W (Credito), T (Categoria)
    
    entries = []
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2): # skip header if any
        cuenta = sheet.cell(row=row_idx, column=6).value # F
        cuenta_nombre = sheet.cell(row=row_idx, column=6).comment
        cuenta_nombre = cuenta_nombre.text.strip() if cuenta_nombre else ""
        
        categoria = sheet.cell(row=row_idx, column=20).value # T
        
        debito = sheet.cell(row=row_idx, column=22).value # V
        debito_comment = sheet.cell(row=row_idx, column=22).comment
        debito_comment = debito_comment.text.strip() if debito_comment else ""
        
        credito = sheet.cell(row=row_idx, column=23).value # W
        credito_comment = sheet.cell(row=row_idx, column=23).comment
        credito_comment = credito_comment.text.strip() if credito_comment else ""
        
        if cuenta or debito or credito:
            entries.append({
                "row": row_idx,
                "cuenta": cuenta,
                "nombre": cuenta_nombre.split("Comentario:\n")[-1].strip() if "Comentario:" in cuenta_nombre else cuenta_nombre,
                "categoria": categoria,
                "debito": debito,
                "debito_logic": debito_comment.split("Comentario:\n")[-1].strip() if "Comentario:" in debito_comment else debito_comment,
                "credito": credito,
                "credito_logic": credito_comment.split("Comentario:\n")[-1].strip() if "Comentario:" in credito_comment else credito_comment
            })

    for entry in entries:
        if entry["nombre"] or entry["debito_logic"] or entry["credito_logic"] or str(entry["debito"]).startswith("=") or str(entry["credito"]).startswith("="):
            print(f"- **Fila {entry['row']} | Cuenta {entry['cuenta']} ({entry['nombre']})**")
            print(f"  - Categoria: {entry['categoria']}")
            if str(entry['debito']).startswith("=") or entry['debito_logic']:
                print(f"  - **Débito:** Formula: `{entry['debito']}` -> *Lógica: {entry['debito_logic']}*")
            if str(entry['credito']).startswith("=") or entry['credito_logic']:
                print(f"  - **Crédito:** Formula: `{entry['credito']}` -> *Lógica: {entry['credito_logic']}*")
            print()
