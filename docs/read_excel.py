import openpyxl
import json

file_path = r"d:\andresContador\Andres-contaduria\archivos extraccion de  data\planoNominalogica.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

output = {}

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    output[sheet_name] = []
    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None or cell.comment is not None:
                cell_data = {
                    "coordinate": cell.coordinate,
                    "value": str(cell.value) if cell.value is not None else None,
                    "comment": cell.comment.text if cell.comment is not None else None
                }
                output[sheet_name].append(cell_data)

print(json.dumps(output, indent=2, ensure_ascii=False))
