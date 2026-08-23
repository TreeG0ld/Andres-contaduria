import openpyxl
from openpyxl.styles import Font, PatternFill
from app.models.config import MapeoPlantilla
from app.models.nomina import LineaNomina, ValorCalculado
from app.models.base import Vinculo, Trabajador, Aportante
from sqlalchemy.orm import Session
import tempfile
import decimal
from datetime import date
import calendar

def get_last_day_of_month(any_date: date) -> date:
    last_day = calendar.monthrange(any_date.year, any_date.month)[1]
    return date(any_date.year, any_date.month, last_day)

def exportar_nomina(db: Session, carga_id: int, period: date, mapeos: list[MapeoPlantilla]) -> str:
    """
    Exporta a un archivo Excel plano de contabilidad.
    Genera 20 filas por cada empleado (según MapeoPlantilla).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ARCHIVO PLANO"
    
    headers = [
        "Tipo de comprobante", "Consecutivo comprobante", "Fecha de elaboración", "Sigla moneda",
        "Tasa de cambio", "Código cuenta contable", "Identificación tercero", "Sucursal",
        "Código producto", "Código de bodega", "Acción", "Cantidad producto", "Prefijo",
        "Consecutivo", "No. cuota", "Fecha vencimiento", "Código impuesto",
        "Código grupo activo fijo", "Código activo fijo", "Descripción",
        "Código centro/subcentro de costos", "Débito", "Crédito", "Observaciones",
        "Base gravable libro compras/ventas", "Base exenta libro compras/ventas", "Mes de cierre"
    ]
    ws.append(headers)
    
    fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font

    mapeos_ordenados = sorted(mapeos, key=lambda m: m.posicion)
    fecha_elaboracion = get_last_day_of_month(period)
    
    # Obtener todas las líneas de la carga
    lineas = db.query(LineaNomina).filter(LineaNomina.carga_id == carga_id).all()
    
    # Consecutivo de comprobante aumenta por cada trabajador
    consecutivo_comprobante = 1
    
    for linea in lineas:
        # Cargar detalles del trabajador y aportante
        vinculo = db.query(Vinculo).filter(Vinculo.id == linea.vinculo_id).first()
        trabajador = db.query(Trabajador).filter(Trabajador.id == vinculo.trabajador_id).first() if vinculo else None
        aportante = db.query(Aportante).filter(Aportante.id == vinculo.aportante_id).first() if vinculo else None
        
        worker_doc = trabajador.numero_documento if trabajador else ""
        clase_gasto = trabajador.clase_gasto if (trabajador and trabajador.clase_gasto) else "51"
        
        # Cargar valores calculados
        valores = db.query(ValorCalculado).filter(ValorCalculado.linea_id == linea.id).all()
        vals_dict = {v.codigo: float(v.valor_original) for v in valores}
        
        for map_row in mapeos_ordenados:
            concept = map_row.codigo_calculo
            acc_tpl = map_row.columna_destino
            side = map_row.formato
            desc = map_row.relleno
            
            amount = vals_dict.get(concept, 0.0)
            
            # Resolver cuenta contable
            account_code = acc_tpl
            if "X" in account_code:
                if clase_gasto == "72" and concept == "auxilio_transporte":
                    account_code = "72072701"
                else:
                    account_code = account_code.replace("X", clase_gasto)
                    
            # Resolver identificación del tercero (Trabajador o Entidad de Seguridad Social para los créditos de aporte)
            third_party = worker_doc
            if map_row.posicion in [8, 10, 12] and aportante:
                if concept == "aporte_pension":
                    third_party = aportante.nit_afp or ""
                elif concept == "aporte_arl":
                    third_party = aportante.nit_arl or ""
                elif concept == "aporte_ccf":
                    third_party = aportante.nit_ccf or ""



                
            # Débito vs Crédito
            debito_val = amount if side == "debito" else 0.0
            credito_val = amount if side == "credito" else 0.0
            
            # Estructurar fila (27 columnas)
            row_data = [""] * 27
            row_data[0] = 8  # Tipo de comprobante (A)
            row_data[1] = consecutivo_comprobante  # Consecutivo aumenta por trabajador (B)
            row_data[2] = fecha_elaboracion.strftime("%Y-%m-%d")  # Fecha (C)
            row_data[3] = "COP"  # Moneda (D)
            row_data[5] = account_code  # Cuenta (F)
            row_data[6] = third_party  # Tercero (G)
            row_data[19] = desc  # Descripción (T)
            row_data[21] = debito_val  # Débito (V)
            row_data[22] = credito_val  # Crédito (W)
            
            ws.append(row_data)
            
        consecutivo_comprobante += 1

            
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp.name)
    return temp.name

